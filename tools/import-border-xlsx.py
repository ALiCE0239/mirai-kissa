#!/usr/bin/env python3
"""data/イベントボーダー一覧.xlsx（シート「イベント別」）から
js/border-rankings-data.js を再生成する。

export-border-xlsx.py の逆方向。Excel を編集用シートとして使い、
行を追記・修正したらこのスクリプトで実行時データを作り直せる。

編集フロー:
  1. PYTHONPATH=./.pylibs python3 tools/export-border-xlsx.py
       → 現行データを data/イベントボーダー一覧.xlsx に書き出し（初回/最新化）
  2. 「イベント別」シートを編集
       - 1イベント＝1行。T10/T50/T100/T500/T1000/T5000 列にボーダーPtを入力
       - 空欄の順位帯は、その帯のデータなしとして扱う（詰めなくてよい）
       - バナー/ユニット/種別/日数/ボーナスも各行に入力
  3. PYTHONPATH=./.pylibs python3 tools/import-border-xlsx.py
       → js/border-rankings-data.js を再生成
  4. ブラウザを強制再読み込み（Cmd+Shift+R）→ git commit / push

使い方:
  PYTHONPATH=./.pylibs python3 tools/import-border-xlsx.py
"""
import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "イベントボーダー一覧.xlsx"
OUT = ROOT / "js" / "border-rankings-data.js"

TIERS = ["10", "50", "100", "500", "1000", "5000"]
PIVOT_SHEET = "イベント別"
DEFAULT_SOURCE = "参考データ/【完全版】プロセカボーダーランキング のコピー"

# 出力する1行のキー順（既存 border-rankings-data.js に合わせる）
ROW_KEYS = ["eventName", "pointsDisplay", "eventType", "banner", "unit", "days", "bonus", "points"]


def points_display(n: int) -> str:
    """points から表示用文字列（例: 約2億3000万P / 約159万P）を生成。"""
    n = int(round(n))
    oku = n // 100_000_000
    man = (n % 100_000_000) // 10_000
    if oku > 0:
        return f"約{oku}億{man}万P" if man > 0 else f"約{oku}億P"
    if man > 0:
        return f"約{man}万P"
    return f"約{n}P"


def clean(v) -> str:
    return "" if v is None else str(v).strip()


def to_points(v):
    if v is None or str(v).strip() == "":
        return None
    s = str(v).replace(",", "").replace("，", "").strip()
    return int(round(float(s)))


def load_existing():
    """既存 JS から EMBEDDED_BORDER_DATA を読む（順序・表示文字列の維持に使用）。"""
    if not OUT.exists():
        return None
    text = OUT.read_text(encoding="utf-8")
    m = re.search(r"const EMBEDDED_BORDER_DATA\s*=\s*(\{.*\});?\s*$", text, re.S)
    if not m:
        return None
    try:
        return json.loads(m.group(1))
    except json.JSONDecodeError:
        return None


def ordered_unique(values: set, prior: list | None) -> list:
    """既存の並び順を優先し、新規値は末尾に五十音（コードポイント）順で追加。"""
    seen = set()
    out = []
    for v in (prior or []):
        if v in values and v not in seen:
            out.append(v)
            seen.add(v)
    for v in sorted(values):
        if v not in seen:
            out.append(v)
            seen.add(v)
    return out


def read_pivot():
    if not XLSX.exists():
        raise SystemExit(
            f"{XLSX} がありません。先に export-border-xlsx.py で書き出してください。"
        )
    wb = load_workbook(XLSX, data_only=True)
    if PIVOT_SHEET not in wb.sheetnames:
        raise SystemExit(f"シート「{PIVOT_SHEET}」が見つかりません（シート: {wb.sheetnames}）")
    ws = wb[PIVOT_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("「イベント別」シートが空です")

    header = [clean(h) for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    for required in ["イベント名"]:
        if required not in idx:
            raise SystemExit(f"ヘッダーに「{required}」列がありません: {header}")

    events = []
    for r in rows[1:]:
        name = clean(r[idx["イベント名"]]) if idx["イベント名"] < len(r) else ""
        if not name:
            continue

        def cell(col_name):
            i = idx.get(col_name)
            return r[i] if (i is not None and i < len(r)) else None

        ev = {
            "eventName": name,
            "banner": clean(cell("バナー")),
            "unit": clean(cell("ユニット")),
            "eventType": clean(cell("種別")),
            "days": clean(cell("日数")),
            "bonus": clean(cell("ボーナス")),
            "tiers": {},
        }
        for t in TIERS:
            pts = to_points(cell(f"T{t}"))
            if pts is not None:
                ev["tiers"][t] = pts
        events.append(ev)
    return events


def build(events, prior):
    # 既存の pointsDisplay を (tier, eventName, points) で引けるように
    prior_display = {}
    if prior:
        for t in TIERS:
            for row in prior.get("ranks", {}).get(t, []):
                prior_display[(t, row.get("eventName"), row.get("points"))] = row.get("pointsDisplay")

    ranks = {t: [] for t in TIERS}
    for ev in events:
        for t, pts in ev["tiers"].items():
            disp = prior_display.get((t, ev["eventName"], pts)) or points_display(pts)
            ranks[t].append({
                "eventName": ev["eventName"],
                "pointsDisplay": disp,
                "eventType": ev["eventType"],
                "banner": ev["banner"],
                "unit": ev["unit"],
                "days": ev["days"],
                "bonus": ev["bonus"],
                "points": pts,
            })
    for t in TIERS:
        ranks[t].sort(key=lambda x: -x["points"])

    banner_vals = {row["banner"] for t in TIERS for row in ranks[t] if row["banner"]}
    unit_vals = {row["unit"] for t in TIERS for row in ranks[t] if row["unit"]}
    banners = ordered_unique(banner_vals, prior.get("banners") if prior else None)
    units = ordered_unique(unit_vals, prior.get("units") if prior else None)
    source = (prior.get("source") if prior else None) or DEFAULT_SOURCE

    # キー順を固定して出力
    ranks_out = {t: [{k: row[k] for k in ROW_KEYS} for row in ranks[t]] for t in TIERS}
    return {"source": source, "ranks": ranks_out, "banners": banners, "units": units}


def main():
    prior = load_existing()
    events = read_pivot()
    data = build(events, prior)
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    OUT.write_text(
        "/** 自動生成: python3 tools/import-border-xlsx.py（元: data/イベントボーダー一覧.xlsx） */\n"
        f"const EMBEDDED_BORDER_DATA = {payload};\n",
        encoding="utf-8",
    )
    total = sum(len(v) for v in data["ranks"].values())
    print(f"再生成完了: {OUT}")
    print(f"  イベント行数: 合計 {total}（" + " / ".join(f"T{t}:{len(data['ranks'][t])}" for t in TIERS) + "）")
    print(f"  バナー {len(data['banners'])}件 / ユニット {len(data['units'])}件")


if __name__ == "__main__":
    main()
