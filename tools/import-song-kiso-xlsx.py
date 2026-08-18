#!/usr/bin/env python3
"""data/楽曲基礎点.xlsx（シート「楽曲基礎点」）から
data/楽曲基礎点.txt を書き戻し、js/song-kiso-data.js と .json を再生成する。

export-song-kiso-xlsx.py の逆方向。txt の「■ リスト」以下だけを Excel の内容で
置き換え、ヘッダー（説明文）はそのまま残す。並び順は Excel の行順を反映する。

編集フロー:
  1. PYTHONPATH=./.pylibs python3 tools/export-song-kiso-xlsx.py
  2. data/楽曲基礎点.xlsx の「楽曲基礎点」シートを編集（曲名/基礎点/ユニット）
  3. PYTHONPATH=./.pylibs python3 tools/import-song-kiso-xlsx.py
  4. ブラウザを強制再読み込み（Cmd+Shift+R）→ git commit / push

使い方:
  PYTHONPATH=./.pylibs python3 tools/import-song-kiso-xlsx.py
"""
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "楽曲基礎点.xlsx"
TXT = ROOT / "data" / "楽曲基礎点.txt"
BUILD = ROOT / "tools" / "build-song-kiso-data.py"
SHEET = "楽曲基礎点"
LIST_MARKER = "■ リスト"


def clean(v) -> str:
    return "" if v is None else str(v).strip()


def read_sheet():
    if not XLSX.exists():
        raise SystemExit(
            f"{XLSX} がありません。先に export-song-kiso-xlsx.py で書き出してください。"
        )
    wb = load_workbook(XLSX, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"シート「{SHEET}」が見つかりません（シート: {wb.sheetnames}）")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("「楽曲基礎点」シートが空です")
    header = [clean(h) for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    for req in ["曲名", "基礎点"]:
        if req not in idx:
            raise SystemExit(f"ヘッダーに「{req}」列がありません: {header}")

    songs = []
    for r in rows[1:]:
        def cell(name):
            i = idx.get(name)
            return r[i] if (i is not None and i < len(r)) else None

        name = clean(cell("曲名"))
        kiso_raw = clean(cell("基礎点"))
        unit = clean(cell("ユニット"))
        if not name:
            continue
        if not kiso_raw:
            print(f"  [警告] 基礎点が空の行をスキップ: {name}")
            continue
        try:
            kiso = int(float(kiso_raw))
        except ValueError:
            print(f"  [警告] 基礎点が数値でない行をスキップ: {name} ({kiso_raw!r})")
            continue
        songs.append((name, kiso, unit))
    return songs


def rewrite_txt(songs):
    """txt の「■ リスト」以下だけを差し替える（ヘッダーは維持）。"""
    lines = TXT.read_text(encoding="utf-8").splitlines()
    head = []
    found = False
    for line in lines:
        head.append(line)
        if line.strip() == LIST_MARKER:
            found = True
            break
    if not found:
        # マーカーが無ければ末尾に付ける
        if head and head[-1].strip():
            head.append("")
        head.append(LIST_MARKER)

    body = [f"{name}\t{kiso}\t{unit}".rstrip("\t") if not unit else f"{name}\t{kiso}\t{unit}"
            for name, kiso, unit in songs]
    TXT.write_text("\n".join(head + body) + "\n", encoding="utf-8")


def main():
    songs = read_sheet()
    if not songs:
        raise SystemExit("Excel から曲を読み取れませんでした")
    rewrite_txt(songs)
    print(f"txt 更新: {TXT}（{len(songs)} 曲）")
    # 実行時データ（js/json）を再生成
    subprocess.run([sys.executable, str(BUILD)], check=True)


if __name__ == "__main__":
    main()
