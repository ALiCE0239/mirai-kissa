#!/usr/bin/env python3
"""data/楽曲基礎点.txt を編集用 Excel(.xlsx) に書き出す。

出力: data/楽曲基礎点.xlsx
  - シート「楽曲基礎点」: No / 曲名 / 基礎点 / ユニット

Excel で編集したあとは import-song-kiso-xlsx.py で txt/js/json を作り直す。

使い方:
  PYTHONPATH=./.pylibs python3 tools/export-song-kiso-xlsx.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "楽曲基礎点.txt"
OUT = ROOT / "data" / "楽曲基礎点.xlsx"
SHEET = "楽曲基礎点"
LIST_MARKER = "■ リスト"

HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def parse_txt():
    songs = []
    in_list = False
    for raw in SRC.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not in_list:
            if line == LIST_MARKER:
                in_list = True
            continue
        if not line or line.startswith("#") or line.startswith("■"):
            continue
        if "\t" in raw:
            parts = [p.strip() for p in raw.split("\t")]
            name = parts[0]
            kiso = parts[1] if len(parts) > 1 else ""
            unit = parts[2] if len(parts) > 2 else ""
        elif "=" in line:
            name, kiso = (p.strip() for p in line.split("=", 1))
            unit = ""
        else:
            continue
        songs.append((name, kiso, unit))
    return songs


def main():
    songs = parse_txt()
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    headers = ["No", "曲名", "基礎点", "ユニット"]
    ws.append(headers)
    for i, (name, kiso, unit) in enumerate(songs, start=1):
        try:
            kiso_val = int(float(kiso))
        except ValueError:
            kiso_val = kiso
        ws.append([i, name, kiso_val, unit])

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for i, w in enumerate([5, 40, 10, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"書き出し完了: {OUT}（{len(songs)} 曲）")


if __name__ == "__main__":
    main()
