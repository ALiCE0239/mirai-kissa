#!/usr/bin/env python3
"""js/border-rankings-data.js の EMBEDDED_BORDER_DATA を Excel(.xlsx) に書き出す。

出力: data/イベントボーダー一覧.xlsx
  - シート「イベント別」: イベント名ごとに各順位帯(T10..T5000)のボーダーPtを横並び
  - シート「T10」「T50」…「T5000」: 順位帯ごとの一覧（ポイント降順）

使い方:
  PYTHONPATH=./.pylibs python3 tools/export-border-xlsx.py
"""
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "js" / "border-rankings-data.js"
OUT = ROOT / "data" / "イベントボーダー一覧.xlsx"

TIERS = ["10", "50", "100", "500", "1000", "5000"]
HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def load_data():
    text = SRC.read_text(encoding="utf-8")
    m = re.search(r"const EMBEDDED_BORDER_DATA\s*=\s*(\{.*\});?\s*$", text, re.S)
    if not m:
        raise SystemExit("EMBEDDED_BORDER_DATA を解析できませんでした")
    return json.loads(m.group(1))


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_pivot(wb, data):
    ranks = data["ranks"]
    # イベント名 -> {tier: points}, メタ情報
    events = {}
    order = []  # T10 のポイント降順を基準にした表示順
    for tier in TIERS:
        for row in ranks.get(tier, []):
            name = row["eventName"]
            if name not in events:
                events[name] = {
                    "banner": row.get("banner", ""),
                    "unit": row.get("unit", ""),
                    "eventType": row.get("eventType", ""),
                    "days": row.get("days", ""),
                    "bonus": row.get("bonus", ""),
                    "points": {},
                }
            events[name]["points"][tier] = row.get("points")

    def sort_key(name):
        p10 = events[name]["points"].get("10")
        # T10 があればポイント降順、無ければ最大のポイントで
        if p10 is not None:
            return (0, -p10)
        best = max([p for p in events[name]["points"].values() if p is not None], default=0)
        return (1, -best)

    order = sorted(events.keys(), key=sort_key)

    ws = wb.active
    ws.title = "イベント別"
    headers = ["No", "イベント名", "バナー", "ユニット", "種別", "日数", "ボーナス"] + [f"T{t}" for t in TIERS]
    ws.append(headers)
    for i, name in enumerate(order, start=1):
        e = events[name]
        row = [i, name, e["banner"], e["unit"], e["eventType"], e["days"], e["bonus"]]
        row += [e["points"].get(t) for t in TIERS]
        ws.append(row)

    ncols = len(headers)
    style_header(ws, ncols)
    autosize(ws, [5, 40, 10, 10, 10, 8, 9] + [15] * len(TIERS))
    # ポイント列を数値書式（桁区切り）
    for col in range(8, ncols + 1):
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = "#,##0"


def build_tier_sheets(wb, data):
    ranks = data["ranks"]
    for tier in TIERS:
        ws = wb.create_sheet(title=f"T{tier}")
        headers = ["No", "イベント名", "ボーダーPt", "表示", "バナー", "ユニット", "種別", "日数", "ボーナス"]
        ws.append(headers)
        for i, row in enumerate(ranks.get(tier, []), start=1):
            ws.append([
                i,
                row.get("eventName", ""),
                row.get("points"),
                row.get("pointsDisplay", ""),
                row.get("banner", ""),
                row.get("unit", ""),
                row.get("eventType", ""),
                row.get("days", ""),
                row.get("bonus", ""),
            ])
        style_header(ws, len(headers))
        autosize(ws, [5, 40, 15, 14, 10, 10, 10, 8, 9])
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=3).number_format = "#,##0"


def main():
    data = load_data()
    wb = Workbook()
    build_pivot(wb, data)
    build_tier_sheets(wb, data)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"書き出し完了: {OUT}")
    print(f"シート: {wb.sheetnames}")


if __name__ == "__main__":
    main()
